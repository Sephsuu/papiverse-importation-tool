import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from openpyxl.drawing.image import Image

data = {
    "content": [
        {
            "branchId": 2,
            "branchName": "Krispy papi - Naic",
            "streetAddress": "129 Governors Drive",
            "barangay": "Malainen Bago",
            "city": "Naic",
            "province": "Cavite",
            "zipCode": 4110,
            "branchStatus": "Open",
            "isInternal": False,
            "isDeleted": False
        },
        {
            "branchId": 3,
            "branchName": "Krispy papi - Silang Maguyam",
            "streetAddress": "Sitio Ibaba",
            "barangay": "Maguyam",
            "city": "Silang",
            "province": "Cavite",
            "zipCode": 4118,
            "branchStatus": "Open",
            "isInternal": False,
            "isDeleted": False
        },
        {
            "branchId": 4,
            "branchName": "Krispy papi - Tanza Paradahan",
            "streetAddress": "Phase 2 Block 3 Lot 1 Westwood Mansion",
            "barangay": "Paradahan 1",
            "city": "Tanza",
            "province": "Cavite",
            "zipCode": 4108,
            "branchStatus": "Open",
            "isInternal": False,
            "isDeleted": False
        },
        {
            "branchId": 15,
            "branchName": "Krispy papi - Sta. Rosa",
            "streetAddress": "Plaza Roseña Food Park Malusak",
            "barangay": "Kanluran",
            "city": "Sta. Rosa",
            "province": "Laguna",
            "zipCode": 4026,
            "branchStatus": "Open",
            "isInternal": False,
            "isDeleted": False
        },
        {
            "branchId": 16,
            "branchName": "Krispy papi - Hugo Perez",
            "streetAddress": "Purok 4",
            "barangay": "Hugo Perez",
            "city": "Trece Martires City",
            "province": "Cavite",
            "zipCode": 4109,
            "branchStatus": "Open",
            "isInternal": False,
            "isDeleted": False
        },
        {
            "branchId": 17,
            "branchName": "Krispy papi - Gentri Manggahan",
            "streetAddress": "Zhang Place, Metro South Village",
            "barangay": "Manggahan",
            "city": "General Trias",
            "province": "Cavite",
            "zipCode": 4107,
            "branchStatus": "Open",
            "isInternal": False,
            "isDeleted": False
        },
        {
            "branchId": 18,
            "branchName": "Krispy papi - Lumil Silang",
            "streetAddress": "Tagaytay Sta. Rosa Rd, Purok 2",
            "barangay": "Lumil",
            "city": "Silang",
            "province": "Cavite",
            "zipCode": 4118,
            "branchStatus": "Open",
            "isInternal": False,
            "isDeleted": False
        },
        {
            "branchId": 19,
            "branchName": "Krispy papi - Dasma Langkaan",
            "streetAddress": "Benaviel Building. 253",
            "barangay": "Langkaan 1",
            "city": "Dasmarinas City",
            "province": "Cavite",
            "zipCode": 4114,
            "branchStatus": "Open",
            "isInternal": False,
            "isDeleted": False
        },
        {
            "branchId": 20,
            "branchName": "Krispy papi - Imus Malagasang",
            "streetAddress": "Greengate Homes Phase 3",
            "barangay": "Maglasan 2A",
            "city": "Imus",
            "province": "Cavite",
            "zipCode": 4103,
            "branchStatus": "Open",
            "isInternal": False,
            "isDeleted": False
        },
        {
            "branchId": 21,
            "branchName": "Krispy papi - Gentri Santiago",
            "streetAddress": "Block 43 Lot 2 Cluster 2 Bella Vista Subd.",
            "barangay": "Santiago",
            "city": "General Trias",
            "province": "Cavite",
            "zipCode": 4107,
            "branchStatus": "Open",
            "isInternal": False,
            "isDeleted": False
        },
        {
            "branchId": 22,
            "branchName": "Krispy papi - Dasma Salawag",
            "streetAddress": "Unit 1A & 1B, Block 1 Lot 78, 80. 82, 84, 86 Armstrong village",
            "barangay": "Salawag",
            "city": "Dasmarinas City",
            "province": "Cavite",
            "zipCode": 4114,
            "branchStatus": "Open",
            "isInternal": False,
            "isDeleted": False
        },
        {
            "branchId": 23,
            "branchName": "Krispy papi - Genstar",
            "streetAddress": "MGSI building",
            "barangay": "Manggahan",
            "city": "General Trias",
            "province": "Cavite",
            "zipCode": 4107,
            "branchStatus": "Open",
            "isInternal": False,
            "isDeleted": False
        },
        {
            "branchId": 1,
            "branchName": "Krispy Papi Main ",
            "streetAddress": "N/A",
            "barangay": "N/A",
            "city": "Silang",
            "province": "Cavite",
            "zipCode": 4109,
            "branchStatus": "ACTIVE",
            "isInternal": False,
            "isDeleted": False
        }
    ],
}

class Generator:
    pass

data = {
    "Branch Name": [item["branchName"].title() for item in data["content"]],
    "Full Address": [f"{item["streetAddress"].upper()}, {item["city"].upper()}, {item["province"].upper()} ({item["zipCode"]}), PHILIPPINES" for item in data["content"]],
    "Status": [item["branchStatus"].title() for item in data["content"]],
    "Branch Type": [
        "Internal Branch" if item["isInternal"] else "External Branch" for item in data["content"]
    ]
}
df = pd.DataFrame(data)
filename = "branches.xlsx"
with pd.ExcelWriter(filename, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, startrow=1, startcol=0)

wb = load_workbook(filename)
ws = wb.active

for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 4 

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = thin_border


ws.merge_cells("A1:D1")
ws.row_dimensions[1].height = 60  
cell = ws["A1"]
cell.value = "Branches of Krispy Papi"
cell.alignment = Alignment(horizontal="center", vertical="center")
img = Image("kp_logo.png")
img.width = 40
img.height = 40
ws.add_image(img, "A1") 
wb.save(filename)